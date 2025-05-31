import akshare as ak
import openpyxl
from openpyxl.utils import get_column_letter
import datetime
import shutil
import pandas as pd
from utils import fin
from pybeans import today
from monthly_info import update_monthly
from daily_info import update_daily

CH_INT = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5}
ANA_PERIODS = ('近1年', '近3年', '近5年', '近10年')
ANA_VALUES = ('年化波动率', '年化夏普比率', '最大回撤')

CODE = 0
NAME = 1
CATE = 2
FOUND_DATE = 4
VALUE = 5
VOLATILITY = 6
SCALE = 7
SALE = 8
FEE = 9
MANAGER = 10
MANAGE_AT = 11
ALL_RETURN = 12
ANA_START = 21


def set_value(ws_cell, df_cell, divider=100):
    if not pd.isna(df_cell):
        ws_cell.value = df_cell / divider


def get_fund_info():
    current_date = datetime.datetime.now().strftime("%Y%m%d")
    file_path = fin['asset_config_path']
    file_path = file_path.replace('.xlsx', f'-{current_date}.xlsx')
    try:
        shutil.copyfile(fin['asset_config_path'], file_path)
    except PermissionError:
        print("文件被占用，请关闭后重试...")
        return
    
    off_exchg_df = ak.fund_open_fund_rank_em(symbol="全部")
    fin.info(f'获取到场外基金数据{len(off_exchg_df)}条')
    on_exchg_df = ak.fund_exchange_rank_em()
    fin.info(f'获取到场内基金数据{len(on_exchg_df)}条')
    money_df = ak.fund_money_rank_em()
    fin.info(f'获取到货币基金数据{len(money_df)}条')
    index_em_df = ak.fund_info_index_em(symbol="全部", indicator="全部")
    fin.info(f'获取到指数基金数据{len(index_em_df)}条')
    #print(df)
    
    VALUE_DATE = None
    
    wb = openpyxl.load_workbook(file_path)
    try:
        sheet = wb['基金']
        data = []
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            #print(row)
            code = row[CODE].value
            name = row[NAME].value
            cate = row[CATE].value
            manager = row[MANAGER].value
            all_return_cell = row[ALL_RETURN]
            
            if not code:
                continue
            
            result = off_exchg_df[off_exchg_df['基金代码'] == code]
            fund_name = ''
            if not result.empty:
                fund_name = result.iloc[0]['基金简称']
            else:
                result = on_exchg_df[on_exchg_df['基金代码'] == code]
                if not result.empty:
                    fund_name = result.iloc[0]['基金简称']
                else:
                    result = index_em_df[index_em_df['基金代码'] == code]
                    if not result.empty:
                        fund_name = result.iloc[0]['基金名称']
                    else:
                        result = money_df[money_df['基金代码'] == code]
                        if not result.empty:
                            pass
                        else:
                            fin.info(f"Line {row_idx}: 未找到基金：{code} {name}")
                            continue
                    
            if fund_name:
                if fund_name != name:
                    fin.error(f"Line {row_idx}: 基金名称发生变化 - 基金：{code}：{name} -> {fund_name}")
                #fin.debug(f"找到基金：{code} {name}")
                
                if not VALUE_DATE:
                    VALUE_DATE = result.iloc[0]['日期']
                    fin.debug(f"净值日期：{VALUE_DATE}")
                    
                old_value = row[VALUE].value
                new_value = result.iloc[0]['单位净值']
                # fin.debug(f"更新基金：{code} {name}：{old_value} -> {new_value}")
                set_value(row[VOLATILITY], (new_value - old_value) / old_value, 1)
                set_value(row[VALUE], new_value, 1)
                set_value(all_return_cell, result.iloc[0]['成立来'])
                set_value(all_return_cell.offset(column=1), result.iloc[0]['近1周'])
                set_value(all_return_cell.offset(column=2), result.iloc[0]['近1月'])
                set_value(all_return_cell.offset(column=3), result.iloc[0]['近3月'])
                set_value(all_return_cell.offset(column=4), result.iloc[0]['近6月'])
                set_value(all_return_cell.offset(column=5), result.iloc[0]['近1年'])
                set_value(all_return_cell.offset(column=6), result.iloc[0]['近2年'])
                set_value(all_return_cell.offset(column=7), result.iloc[0]['近3年'])
                set_value(all_return_cell.offset(column=8), result.iloc[0]['今年来'])
                
                try:
                    fee_df = ak.fund_individual_detail_info_xq(symbol=code)
                    last_sell_rule = fee_df[fee_df['费用类型'] == '卖出规则'].iloc[-1]
                    sale_fee = ''
                    if last_sell_rule['费用'] > 0:
                        sale_fee = f'({last_sell_rule['费用']})'
                    row[SALE].value = f'{last_sell_rule['条件或名称'].replace("持有期限", "").replace("<=", "")}{sale_fee}'
                    other_fees_df = fee_df[fee_df['费用类型'] == '其他费用']
                    formula = "=(" + "+".join([str(fee) for fee in other_fees_df['费用']]) + ")/100"
                    row[FEE].value = formula
                except Exception as ex:
                    #raise ex
                    fin.info(f"Line {row_idx}: 交易信息获取失败 - 基金：{code} {name}, {str(ex)}")
                    continue
                
                try:
                    basic_df = ak.fund_individual_basic_info_xq(symbol=code)
                    row[FOUND_DATE].value = basic_df[basic_df['item']=='成立时间'].iloc[0]['value']
                    new_manager = basic_df[basic_df['item']=='基金经理'].iloc[0]['value']
                    if new_manager != manager:
                        row[MANAGER].value = new_manager
                        if not (set(manager.split(' ')) & set(new_manager.split(' '))): ## 没有交集说明是不是加入搭档，更新管理时间
                            row[MANAGE_AT].value = today() 
                        if manager:
                            fin.error(f"Line {row_idx}: 基金经理发生变化 - 基金：{code} {name}：{manager} -> {new_manager}")
                        
                    scale = basic_df[basic_df['item']=='最新规模'].iloc[0]['value']
                    scale = scale.replace("亿", "")
                    if scale.endswith("万"):
                        scale = float(scale.replace("万", "")) / 10000
                    row[SCALE].value = float(scale)
                except Exception as ex:
                    #raise ex
                    fin.info(f"Line {row_idx}: 基础信息获取失败 - 基金：{code} {name}， {str(ex)}")
                    continue
                
                if cate not in ('长债', '中债', '短债') and 'ETF' not in name and '指数' not in name and '联接' not in name:
                    try:
                        ana_df = ak.fund_individual_analysis_xq(symbol=code)
                        for apn, apv in enumerate(ANA_PERIODS):
                            found_row = ana_df[ana_df['周期'] == apv]
                            if not found_row.empty:
                                for avn, avv in enumerate(ANA_VALUES):
                                    row[ANA_START + apn*3 + avn].value = found_row[avv].values[0]
                    except Exception as ex:
                        #raise ex
                        fin.info(f"Line {row_idx}: 分析信息获取失败 - 基金：{code} {name}， {str(ex)}")
                        continue
            else:
                try:
                    hist_df = ak.fund_open_fund_info_em(symbol=code, indicator="累计净值走势")
                    old_value = row[VALUE].value
                    new_value = hist_df.iloc[-1]['累计净值']
                    set_value(row[VOLATILITY], (new_value - old_value) / old_value, 1)
                    set_value(row[VALUE], hist_df.iloc[-1]['累计净值'], 1)
                    if not VALUE_DATE:
                        VALUE_DATE = hist_df.iloc[-1]['净值日期']
                        fin.debug(f"净值日期：{VALUE_DATE}")
                    hist_df = ak.fund_open_fund_info_em(symbol=code, indicator="累计收益率走势", period="成立来")
                    set_value(all_return_cell, hist_df.iloc[-1]['累计收益率'])
                except Exception as ex:
                    #raise ex
                    fin.info(f"Line {row_idx}: 收益信息获取失败 - 基金：{code} {name}， {str(ex)}")
                    continue
        if VALUE_DATE:
            sheet[f'{get_column_letter(VALUE+1)}1'] = f'净值({VALUE_DATE})'
        wb.save(file_path)
    finally:
        wb.close()
        
    fin.info(f"> 更新月度宏观数据")
    update_monthly(file_path)
    
    fin.info(f"> 更新每日高频数据")
    update_daily(file_path)

if __name__ == "__main__":
    get_fund_info()
    input('Press any key to quit')